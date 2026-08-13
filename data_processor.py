"""
data_processor.py — adaptado para SharePoint/OneDrive
------------------------------------------------------
Los dos archivos delta (MODIFICACIONES_SABADOS y MODIFICACIONES_PERSONAL) se leen y
escriben en SharePoint mediante graph_client.py.

El archivo maestro (TURNOS SABADOS.xlsx) sigue siendo de SOLO LECTURA: se descarga
por Graph API solo para cargar los turnos base, igual que en el despliegue anterior.

Cuando se ejecuta en local sin credenciales Azure configuradas, usa rutas de archivo
locales como fallback (comportamiento original).
"""

import pandas as pd
import streamlit as st
import openpyxl
import re
import datetime
import os
import io

# ---------------------------------------------------------------------------
# Importación condicional del cliente de Graph API
# ---------------------------------------------------------------------------
try:
    import graph_client as gc
except Exception:
    gc = None

def _use_sharepoint() -> bool:
    try:
        if gc is not None:
            return gc.is_sharepoint_configured()
    except Exception:
        pass
    return False

MONTH_MAP = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'mrazo': 3,
    'abril': 4, 'mayo': 5, 'junio': 6, 'juniode': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9,
    'octubre': 10, 'noviembre': 11, 'diciembre': 12
}

MONTH_NAMES_SP = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
    5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}

# ---------------------------------------------------------------------------
# Helpers de acceso a los archivos delta (SharePoint o local)
# ---------------------------------------------------------------------------

def _get_modifications_path(main_excel_path):
    return os.path.join(os.path.dirname(main_excel_path), "MODIFICACIONES_SABADOS.xlsx")


def _load_wb_delta(file_key: str, local_path: str, sheet_title: str, cols: list):
    """
    Carga el workbook del archivo delta.
    - En SharePoint: descarga a BytesIO.
    - En local: abre desde disco; si no existe, lo crea.
    Devuelve (wb, ws).
    """
    if _use_sharepoint():
        try:
            buf = gc.download_excel(file_key)
            wb = openpyxl.load_workbook(buf)
        except Exception:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            ws.append(cols)
            return wb, ws
            
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
        else:
            ws = wb.create_sheet(title=sheet_title)
            ws.append(cols)
        return wb, ws
    else:
        if not os.path.exists(local_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            ws.append(cols)
            wb.save(local_path)
            wb.close()
        wb = openpyxl.load_workbook(local_path)
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
        else:
            ws = wb.create_sheet(title=sheet_title)
            ws.append(cols)
        return wb, ws


def _save_wb_delta(wb: openpyxl.Workbook, file_key: str, local_path: str):
    """
    Guarda el workbook del archivo delta.
    - En SharePoint: serializa a BytesIO y sube.
    - En local: guarda en disco.
    """
    if _use_sharepoint():
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        gc.upload_excel(file_key, buf)
        wb.close()
    else:
        wb.save(local_path)
        wb.close()


def _read_delta_df(file_key: str, local_path: str, sheet_title: str, cols: list) -> pd.DataFrame:
    """
    Lee el DataFrame del archivo delta (SharePoint o local).
    Si el remoto en SharePoint está vacío pero el archivo local del repositorio tiene registros,
    utiliza el local y lo autorrestaura automáticamente en SharePoint.
    """
    if _use_sharepoint():
        df_remote = pd.DataFrame(columns=cols)
        try:
            buf = gc.download_excel(file_key)
            df_temp = pd.read_excel(buf, sheet_name=sheet_title)
            for c in cols:
                if c not in df_temp.columns:
                    df_temp[c] = None
            df_remote = df_temp[cols].copy()
        except Exception:
            pass

        if df_remote.empty and os.path.exists(local_path):
            try:
                df_local = pd.read_excel(local_path, sheet_name=sheet_title)
                for c in cols:
                    if c not in df_local.columns:
                        df_local[c] = None
                df_local_clean = df_local[cols].copy()
                if not df_local_clean.empty:
                    try:
                        wb_local = openpyxl.load_workbook(local_path)
                        _save_wb_delta(wb_local, file_key, local_path)
                    except Exception as up_err:
                        print(f"Error auto-restaurando delta en SharePoint: {up_err}")
                    return df_local_clean
            except Exception:
                pass

        return df_remote
    else:
        if not os.path.exists(local_path):
            return pd.DataFrame(columns=cols)
        try:
            df = pd.read_excel(local_path, sheet_name=sheet_title)
            for c in cols:
                if c not in df.columns:
                    df[c] = None
            return df[cols].copy()
        except Exception:
            return pd.DataFrame(columns=cols)


# ---------------------------------------------------------------------------
# Parsers de fechas (sin cambios)
# ---------------------------------------------------------------------------

def clean_month_year(header_str):
    if not isinstance(header_str, str):
        return None, None, None
    header_str = header_str.strip().lower()
    header_str = header_str.replace("mes:", "").replace(" de ", " ").strip()
    header_str = re.sub(r'\s+', ' ', header_str)
    parts = header_str.split(' ')
    if len(parts) >= 2:
        month_word = parts[0]
        year_str = parts[-1]
        month = MONTH_MAP.get(month_word, None)
        try:
            year = int(year_str)
            return month, year, month_word.upper()
        except ValueError:
            return None, None, None
    return None, None, None


def parse_date_cell(cell, expected_month, expected_year):
    if pd.isna(cell):
        return None
    if isinstance(cell, datetime.datetime):
        if cell.month == expected_month and cell.year == expected_year:
            return cell.date()
        if cell.day == expected_month and cell.year == expected_year:
            try:
                return datetime.date(expected_year, expected_month, cell.month)
            except Exception:
                pass
        return cell.date()
    cell_str = str(cell).strip().lower()
    if not cell_str or cell_str == '\xa0':
        return None
    match_slash = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', cell_str)
    if match_slash:
        d, m, y = int(match_slash.group(1)), int(match_slash.group(2)), int(match_slash.group(3))
        if m != expected_month and d == expected_month:
            d, m = m, d
        try:
            return datetime.date(y, m, d)
        except ValueError:
            pass
    cell_str = re.sub(r'\s+', ' ', cell_str)
    day_match = re.search(r'\b(\d{1,2})\b', cell_str)
    if day_match:
        day = int(day_match.group(1))
        month = expected_month
        for m_name, m_val in MONTH_MAP.items():
            if m_name in cell_str:
                month = m_val
                break
        year = expected_year
        year_match = re.search(r'\b(202\d)\b', cell_str)
        if year_match:
            year = int(year_match.group(1))
        try:
            return datetime.date(year, month, day)
        except ValueError:
            pass
    return None


def parse_flat_date(cell):
    if pd.isna(cell):
        return None
    if isinstance(cell, datetime.datetime):
        return cell.date()
    if isinstance(cell, datetime.date):
        return cell
    cell_str = str(cell).strip()
    if not cell_str or cell_str == '\xa0':
        return None
    match_iso = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})', cell_str)
    if match_iso:
        try:
            return datetime.date(int(match_iso.group(1)), int(match_iso.group(2)), int(match_iso.group(3)))
        except ValueError:
            pass
    match_slash = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})', cell_str)
    if match_slash:
        try:
            return datetime.date(int(match_slash.group(3)), int(match_slash.group(2)), int(match_slash.group(1)))
        except ValueError:
            pass
    try:
        return pd.to_datetime(cell_str).date()
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# MODIFICACIONES_PERSONAL
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def load_personal_modifications(excel_path):
    # Descontinuado: El personal se lee directamente desde SharePoint BD PERSONAL (CONSOLIDADO 2026.xlsx)
    return pd.DataFrame()


# ---------------------------------------------------------------------------
# PLAN DE CONTINGENCIA: Personal Manual (Guardado en MODIFICACIONES_SABADOS.xlsx)
# ---------------------------------------------------------------------------

_COLS_MANUAL_PERSONAL = ['ID', 'CEDULA', 'NOMBRES_Y_APELLIDOS', 'CARGO', 'CELULAR', 'SEDE_CECO', 'STATUS', 'OBSERVACIONES', 'TIMESTAMP']
_SHEET_MANUAL_PERSONAL = 'PERSONAL_MANUAL'


@st.cache_data(show_spinner=False)
def load_manual_supernumeraries(excel_path):
    """
    Carga los médicos registrados manualmente por contingencia desde la pestaña 'PERSONAL_MANUAL'
    del archivo delta MODIFICACIONES_SABADOS.xlsx.
    """
    local_path = _get_modifications_path(excel_path)
    df = _read_delta_df(_KEY_SABADOS, local_path, _SHEET_MANUAL_PERSONAL, _COLS_MANUAL_PERSONAL)

    if df.empty:
        return df

    df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
    df['CEDULA'] = df['CEDULA'].astype(str).str.strip().apply(
        lambda x: str(int(float(x))) if str(x).replace('.0', '').isdigit() else str(x)
    )
    df['NOMBRES Y APELLIDOS'] = df['NOMBRES_Y_APELLIDOS'].fillna('').astype(str).str.strip().str.upper()
    df['CARGO'] = df['CARGO'].fillna('MEDICO GENERAL SUPERNUMERARIO').astype(str).str.strip().str.upper()
    df['CELULAR'] = df['CELULAR'].fillna('').astype(str).str.strip()
    df['SEDE / CECO'] = df['SEDE_CECO'].fillna('SUPERNUMERARIOS').astype(str).str.strip().str.upper()
    df['STATUS'] = df['STATUS'].fillna('ACTIVO').astype(str).str.strip().str.upper()
    df['OBSERVACIONES'] = df['OBSERVACIONES'].fillna('').astype(str).str.strip()
    
    return df[df['STATUS'] == 'ACTIVO'].sort_values(by='ID')


def save_manual_supernumerary(excel_path, doc_data):
    """
    Guarda un médico registrado manualmente por contingencia en la pestaña 'PERSONAL_MANUAL'
    de MODIFICACIONES_SABADOS.xlsx (sincronizado con SharePoint).
    """
    local_path = _get_modifications_path(excel_path)
    wb, ws = _load_wb_delta(_KEY_SABADOS, local_path, _SHEET_MANUAL_PERSONAL, _COLS_MANUAL_PERSONAL)

    # Verificar encabezados
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if 'OBSERVACIONES' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'OBSERVACIONES'
    if 'TIMESTAMP' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'TIMESTAMP'

    max_id = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except ValueError:
                pass
    next_id = max_id + 1

    ws.append([
        next_id,
        str(doc_data.get('cedula', '')).strip(),
        str(doc_data.get('nombres_y_apellidos', '')).strip().upper(),
        str(doc_data.get('cargo', 'MEDICO GENERAL SUPERNUMERARIO')).strip().upper(),
        str(doc_data.get('celular', '')).strip(),
        str(doc_data.get('sede_ceco', 'SUPERNUMERARIOS')).strip().upper(),
        'ACTIVO',
        str(doc_data.get('observaciones', 'Registro manual contingencia')).strip(),
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    _save_wb_delta(wb, _KEY_SABADOS, local_path)
    load_manual_supernumeraries.clear()
    load_supernumeraries.clear()
    return next_id


def deactivate_manual_supernumerary(excel_path, cedula):
    """
    Marca un médico registrado manualmente como INACTIVO en la pestaña 'PERSONAL_MANUAL'.
    """
    local_path = _get_modifications_path(excel_path)
    wb, ws = _load_wb_delta(_KEY_SABADOS, local_path, _SHEET_MANUAL_PERSONAL, _COLS_MANUAL_PERSONAL)
    
    cedula_str = str(cedula).strip()
    header = [str(ws.cell(row=1, column=c).value).upper() for c in range(1, ws.max_column + 1)]
    ced_idx = header.index('CEDULA') + 1 if 'CEDULA' in header else 2
    status_idx = header.index('STATUS') + 1 if 'STATUS' in header else 7
    
    modified = False
    for r in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row=r, column=ced_idx).value).strip()
        if cell_val.replace('.0', '') == cedula_str.replace('.0', ''):
            ws.cell(row=r, column=status_idx).value = 'INACTIVO'
            modified = True
            
    if modified:
        _save_wb_delta(wb, _KEY_SABADOS, local_path)
        load_manual_supernumeraries.clear()
        load_supernumeraries.clear()
    return modified


# ---------------------------------------------------------------------------
# MODIFICACIONES_SABADOS
# ---------------------------------------------------------------------------

_COLS_SABADOS = ['ID', 'SHEET', 'DATE', 'ORIGINAL_NAME', 'NEW_NAME',
                 'ROW', 'COL', 'TYPE', 'OBSERVACIONES', 'CLASIFICACION', 'TIMESTAMP']
_SHEET_SABADOS = 'MODIFICACIONES'
_KEY_SABADOS   = 'modificaciones_sabados'


@st.cache_data(show_spinner=False)
def load_modifications(excel_path):
    local_path = _get_modifications_path(excel_path)
    df = _read_delta_df(_KEY_SABADOS, local_path, _SHEET_SABADOS, _COLS_SABADOS)

    if df.empty:
        return df

    df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
    df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce').dt.date
    df['SHEET'] = df['SHEET'].astype(str).str.strip()
    df['ORIGINAL_NAME'] = df['ORIGINAL_NAME'].fillna('').astype(str).str.strip().str.upper()
    df['NEW_NAME'] = df['NEW_NAME'].fillna('').astype(str).str.strip().str.upper()
    df['ROW'] = pd.to_numeric(df['ROW'], errors='coerce').fillna(0).astype(int)
    df['COL'] = pd.to_numeric(df['COL'], errors='coerce').fillna(0).astype(int)
    df['TYPE'] = df['TYPE'].astype(str).str.strip().str.upper()
    df['OBSERVACIONES'] = df['OBSERVACIONES'].fillna('').astype(str).str.strip()
    df['CLASIFICACION'] = df['CLASIFICACION'].fillna('Secuencia Normal').astype(str).str.strip()
    return df.sort_values(by='ID')


def save_modification(excel_path, mod_data):
    local_path = _get_modifications_path(excel_path)
    wb, ws = _load_wb_delta(_KEY_SABADOS, local_path, _SHEET_SABADOS, _COLS_SABADOS)

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if 'OBSERVACIONES' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'OBSERVACIONES'
        header.append('OBSERVACIONES')
    if 'CLASIFICACION' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'CLASIFICACION'
        header.append('CLASIFICACION')

    if 'TIMESTAMP' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'TIMESTAMP'
        header.append('TIMESTAMP')

    max_id = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except ValueError:
                pass
    next_id = max_id + 1

    date_val = mod_data['date']
    if isinstance(date_val, (datetime.date, datetime.datetime)):
        date_val = date_val.strftime('%Y-%m-%d')

    val_map = {
        'ID': next_id,
        'SHEET': mod_data['sheet'],
        'DATE': date_val,
        'ORIGINAL_NAME': mod_data.get('original_name', '').strip().upper(),
        'NEW_NAME': mod_data.get('new_name', '').strip().upper(),
        'ROW': mod_data.get('row', 0),
        'COL': mod_data.get('col', 0),
        'TYPE': mod_data['type'].strip().upper(),
        'OBSERVACIONES': str(mod_data.get('observaciones', '')).strip(),
        'CLASIFICACION': mod_data.get('clasificacion', 'Secuencia Normal').strip(),
        'TIMESTAMP': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    row_data = [val_map.get(col_name, '') for col_name in header]
    ws.append(row_data)

    _save_wb_delta(wb, _KEY_SABADOS, local_path)
    load_modifications.clear()
    return next_id


def save_modifications_batch(excel_path, mods_list):
    if not mods_list:
        return True
    local_path = _get_modifications_path(excel_path)
    wb, ws = _load_wb_delta(_KEY_SABADOS, local_path, _SHEET_SABADOS, _COLS_SABADOS)

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if 'OBSERVACIONES' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'OBSERVACIONES'
        header.append('OBSERVACIONES')
    if 'CLASIFICACION' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'CLASIFICACION'
        header.append('CLASIFICACION')
    if 'TIMESTAMP' not in header:
        ws.cell(row=1, column=ws.max_column + 1).value = 'TIMESTAMP'
        header.append('TIMESTAMP')

    max_id = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except ValueError:
                pass
    next_id = max_id + 1

    for mod_data in mods_list:
        date_val = mod_data['date']
        if isinstance(date_val, (datetime.date, datetime.datetime)):
            date_val = date_val.strftime('%Y-%m-%d')

        val_map = {
            'ID': next_id,
            'SHEET': mod_data['sheet'],
            'DATE': date_val,
            'ORIGINAL_NAME': mod_data.get('original_name', '').strip().upper(),
            'NEW_NAME': mod_data.get('new_name', '').strip().upper(),
            'ROW': mod_data.get('row', 0),
            'COL': mod_data.get('col', 0),
            'TYPE': mod_data['type'].strip().upper(),
            'OBSERVACIONES': str(mod_data.get('observaciones', '')).strip(),
            'CLASIFICACION': mod_data.get('clasificacion', 'Secuencia Normal').strip(),
            'TIMESTAMP': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        row_data = [val_map.get(col_name, '') for col_name in header]
        ws.append(row_data)
        next_id += 1

    _save_wb_delta(wb, _KEY_SABADOS, local_path)
    load_modifications.clear()
    return True


# ---------------------------------------------------------------------------
# Carga del Excel maestro (solo lectura — igual que el despliegue anterior)
# ---------------------------------------------------------------------------

@st.cache_resource(ttl=600, show_spinner=False)
def _open_master_excel(excel_path):
    """
    Devuelve un pd.ExcelFile del Excel maestro (TURNOS SABADOS.xlsx).
    En SharePoint: usa una caché en disco local (TURNOS_SABADOS_cached.xlsx).
      1. Obtiene la metadata (lastModifiedDateTime y size) del archivo remoto.
      2. Compara con la metadata local guardada en TURNOS_SABADOS_cached_meta.txt.
      3. Si coincide y el archivo local existe, lee el archivo local directamente (ahorra descargar 4.4MB).
      4. Si no coincide, descarga el archivo de SharePoint, actualiza la caché local y lo lee.
    En local: abre desde la ruta directamente.
    """
    if _use_sharepoint():
        cache_file = "TURNOS_SABADOS_cached.xlsx"
        meta_file = "TURNOS_SABADOS_cached_meta.txt"
        
        try:
            # 1. Obtener metadatos remotos de forma rápida
            remote_meta = gc.get_file_metadata("turnos_sabados")
            remote_last_mod = remote_meta.get("lastModifiedDateTime", "")
            remote_size = remote_meta.get("size", 0)
            
            # 2. Comprobar si tenemos el archivo guardado y coincide
            use_cached = False
            if os.path.exists(cache_file) and os.path.exists(meta_file):
                try:
                    with open(meta_file, "r", encoding="utf-8") as f:
                        saved_meta = f.read().strip().split("|")
                    if len(saved_meta) == 2:
                        saved_last_mod, saved_size = saved_meta[0], int(saved_meta[1])
                        if saved_last_mod == remote_last_mod and saved_size == remote_size:
                            use_cached = True
                except Exception:
                    pass
            
            # 3. Si coincide, usar el archivo local
            if use_cached:
                return pd.ExcelFile(cache_file)
                
            # 4. Si no coincide, descargar el archivo completo
            buf = gc.download_excel("turnos_sabados")
            # Guardar en disco local
            try:
                with open(cache_file, "wb") as f:
                    f.write(buf.getvalue())
                with open(meta_file, "w", encoding="utf-8") as f:
                    f.write(f"{remote_last_mod}|{remote_size}")
            except Exception as cache_err:
                print(f"Error escribiendo en caché de disco local: {cache_err}")
                
            buf.seek(0)
            return pd.ExcelFile(buf)
            
        except Exception as e:
            # Fallback offline si la conexión a SharePoint falla
            if os.path.exists(cache_file):
                try:
                    return pd.ExcelFile(cache_file)
                except Exception:
                    pass
            raise FileNotFoundError(f"No se pudo descargar ni recuperar de la caché local el Excel maestro desde SharePoint: {e}")
    else:
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"El archivo Excel no existe en: {excel_path}")
        return pd.ExcelFile(excel_path)



@st.cache_data(ttl=3600, show_spinner=False)
def _get_base_shifts_df(excel_path):
    xl = _open_master_excel(excel_path)

    months = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
              'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

    candidate_sheets = []
    for s in xl.sheet_names:
        s_upper = s.upper()
        if s_upper in ["SABADOS 2025", "SABADOS 2026"]:
            candidate_sheets.append(s)
        elif any(m in s_upper for m in months) and re.search(r'\b(202\d)\b', s_upper):
            if not s_upper.startswith("TABLA") and "EXTRA" not in s_upper:
                candidate_sheets.append(s)

    all_shifts = []
    errors = []

    for sheet_name in candidate_sheets:
        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
        if len(df) == 0:
            continue

        header_row_idx = None
        sabado_col_idx = None
        super_col_idx = None

        for r_idx in range(min(5, len(df))):
            row_vals = [str(val).strip().upper() for val in df.iloc[r_idx] if pd.notna(val)]
            if 'SABADO' in row_vals and 'SUPERNUMERARIO' in row_vals:
                header_row_idx = r_idx
                for c_idx, val in enumerate(df.iloc[r_idx]):
                    if pd.notna(val):
                        val_str = str(val).strip().upper()
                        if val_str == 'SABADO':
                            sabado_col_idx = c_idx
                        elif val_str == 'SUPERNUMERARIO':
                            super_col_idx = c_idx
                break

        if header_row_idx is not None and sabado_col_idx is not None and super_col_idx is not None:
            month_header = sheet_name.split(' ')[0].upper()
            for idx in range(header_row_idx + 1, len(df)):
                row = df.iloc[idx]
                date_cell = row[sabado_col_idx]
                name_cell = row[super_col_idx]
                if pd.notna(date_cell) and pd.notna(name_cell):
                    d = parse_flat_date(date_cell)
                    name_str = str(name_cell).strip().upper()
                    name_str = re.sub(r'\s+', ' ', name_str)
                    if name_str and name_str not in ["VALENCIA", "MES:", "SABADOS", "TOTAL", "CANTIDAD"]:
                        if not name_str.replace('.', '', 1).isdigit() and len(name_str) > 3:
                            if d:
                                all_shifts.append({
                                    'Sheet': sheet_name, 'Month_Header': month_header,
                                    'Date': d, 'Year': d.year, 'Month': d.month,
                                    'Supernumerary': name_str,
                                    'Excel_Row': int(idx + 1),
                                    'Excel_Col': int(super_col_idx + 1),
                                    'Header_Row': int(header_row_idx + 1)
                                })
                            else:
                                errors.append(f"No se pudo parsear la fecha en fila {idx+1} de {sheet_name}")

        elif sheet_name.upper() in ["SABADOS 2025", "SABADOS 2026"]:
            current_month_num = None
            current_year = None
            current_month_name = None
            date_cols = {}
            header_row_idx = None
            i = 0
            n_rows = len(df)
            while i < n_rows:
                row = df.iloc[i]
                is_header = False
                header_val = None
                for val in row:
                    if isinstance(val, str) and "MES:" in val:
                        is_header = True
                        header_val = val
                        break
                if is_header:
                    current_month_num, current_year, current_month_name = clean_month_year(header_val)
                    header_row_idx = i
                    i += 1
                    if i >= n_rows:
                        break
                    date_row = df.iloc[i]
                    date_cols = {}
                    for col_idx, val in enumerate(date_row):
                        if pd.notna(val):
                            d = parse_date_cell(val, current_month_num, current_year)
                            if d:
                                date_cols[col_idx] = d
                            else:
                                val_str = str(val).strip()
                                if val_str and val_str != '\xa0':
                                    errors.append(f"No se pudo parsear fecha '{val_str}' fila {i+1} de {sheet_name}")
                else:
                    if current_month_num is not None:
                        for col_idx, d in date_cols.items():
                            if col_idx < len(row):
                                name = row[col_idx]
                                if pd.notna(name):
                                    name_str = str(name).strip()
                                    name_str = re.sub(r'\s+', ' ', name_str).upper()
                                    if name_str and name_str not in ["VALENCIA", "MES:", "SABADOS", "TOTAL", "CANTIDAD"]:
                                        if not name_str.replace('.', '', 1).isdigit() and len(name_str) > 3:
                                            all_shifts.append({
                                                'Sheet': sheet_name, 'Month_Header': current_month_name,
                                                'Date': d, 'Year': d.year, 'Month': d.month,
                                                'Supernumerary': name_str,
                                                'Excel_Row': int(i + 1),
                                                'Excel_Col': int(col_idx + 1),
                                                'Header_Row': int(header_row_idx + 1)
                                            })
                i += 1

    df_shifts = pd.DataFrame(all_shifts) if all_shifts else pd.DataFrame(columns=[
        'Sheet', 'Month_Header', 'Date', 'Year', 'Month', 'Supernumerary',
        'Excel_Row', 'Excel_Col', 'Header_Row'
    ])
    df_shifts['Excel_Row'] = df_shifts['Excel_Row'].astype(int)
    df_shifts['Excel_Col'] = df_shifts['Excel_Col'].astype(int)
    df_shifts['Observation'] = ''
    df_shifts['Classification'] = 'Secuencia Normal'
    return df_shifts, errors

def load_data(excel_path):
    """
    Lee los turnos del Excel maestro y aplica las modificaciones delta en memoria.
    """
    df_shifts_base, errors_base = _get_base_shifts_df(excel_path)
    df_shifts = df_shifts_base.copy()
    errors = list(errors_base)

    # Aplicar modificaciones delta en memoria
    df_mods = load_modifications(excel_path)

    for _, mod in df_mods.iterrows():
        m_type = mod['TYPE']
        sheet  = mod['SHEET']
        m_date = mod['DATE']
        orig   = mod['ORIGINAL_NAME']
        new    = mod['NEW_NAME']
        row    = mod['ROW']
        col    = mod['COL']

        if m_type == 'REEMPLAZAR':
            obs    = mod.get('OBSERVACIONES', '')
            clasif = mod.get('CLASIFICACION', 'Secuencia Normal')
            if row > 0 and col > 0:
                mask = ((df_shifts['Sheet'] == sheet) &
                        (df_shifts['Excel_Row'] == row) &
                        (df_shifts['Excel_Col'] == col))
                if mask.any():
                    df_shifts.loc[mask, 'Supernumerary'] = new
                    df_shifts.loc[mask, 'Observation']   = obs
                    df_shifts.loc[mask, 'Classification'] = clasif
                else:
                    mask_fb = ((df_shifts['Sheet'] == sheet) &
                               (df_shifts['Date'] == m_date) &
                               (df_shifts['Supernumerary'] == orig))
                    df_shifts.loc[mask_fb, 'Supernumerary'] = new
                    df_shifts.loc[mask_fb, 'Observation']   = obs
                    df_shifts.loc[mask_fb, 'Classification'] = clasif
            else:
                mask = ((df_shifts['Sheet'] == sheet) &
                        (df_shifts['Date'] == m_date) &
                        (df_shifts['Supernumerary'] == orig))
                df_shifts.loc[mask, 'Supernumerary'] = new
                df_shifts.loc[mask, 'Observation']   = obs
                df_shifts.loc[mask, 'Classification'] = clasif

        elif m_type == 'ELIMINAR':
            if row > 0 and col > 0:
                df_shifts = df_shifts[~((df_shifts['Sheet'] == sheet) &
                                        (df_shifts['Excel_Row'] == row) &
                                        (df_shifts['Excel_Col'] == col))]
            else:
                df_shifts = df_shifts[~((df_shifts['Sheet'] == sheet) &
                                        (df_shifts['Date'] == m_date) &
                                        (df_shifts['Supernumerary'] == orig))]

        elif m_type == 'AGREGAR':
            mask = (df_shifts['Date'] == m_date) & (df_shifts['Supernumerary'] == new)
            if mask.any():
                df_shifts.loc[mask, 'Observation'] = mod.get('OBSERVACIONES', '')
                df_shifts.loc[mask, 'Classification'] = mod.get('CLASIFICACION', 'Secuencia Normal')
            else:
                new_row = {
                    'Sheet': sheet,
                    'Month_Header': MONTH_NAMES_SP.get(m_date.month, 'EXTRA'),
                    'Date': m_date, 'Year': m_date.year, 'Month': m_date.month,
                    'Supernumerary': new,
                    'Excel_Row': 0, 'Excel_Col': 0, 'Header_Row': 0,
                    'Observation': mod.get('OBSERVACIONES', ''),
                    'Classification': mod.get('CLASIFICACION', 'Secuencia Normal')
                }
                df_shifts = pd.concat([df_shifts, pd.DataFrame([new_row])], ignore_index=True)

    # Limpieza final de seguridad contra race conditions (múltiples usuarios) o errores en el maestro
    if not df_shifts.empty:
        df_shifts = df_shifts.drop_duplicates(subset=['Date', 'Supernumerary'], keep='last')

    return df_shifts, errors


# ---------------------------------------------------------------------------
# Supernumerarios (solo lectura del maestro + delta personal)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Supernumerarios (Lectura desde SharePoint: CONSOLIDADO 2026.xlsx - BD PERSONAL)
# ---------------------------------------------------------------------------

@st.cache_resource(ttl=15, show_spinner=False)
def _open_consolidado_personal(excel_path):
    """
    Obtiene pd.ExcelFile para CONSOLIDADO 2026.xlsx desde SharePoint o caché/fallback local.
    Sincronizado en tiempo real (TTL=15s).
    """
    cache_file = "CONSOLIDADO_2026_cached.xlsx"
    meta_file = "CONSOLIDADO_2026_cached_meta.txt"
    force_refresh = st.session_state.pop("force_refresh_personal", False)

    if force_refresh:
        if os.path.exists(cache_file):
            try:
                os.remove(cache_file)
            except Exception:
                pass
        if os.path.exists(meta_file):
            try:
                os.remove(meta_file)
            except Exception:
                pass

    if _use_sharepoint():
        try:
            remote_meta = gc.get_file_metadata("consolidado_personal")
            remote_last_mod = remote_meta.get("lastModifiedDateTime", "")
            remote_size = remote_meta.get("size", 0)
            
            use_cached = False
            if not force_refresh and os.path.exists(cache_file) and os.path.exists(meta_file):
                try:
                    with open(meta_file, "r", encoding="utf-8") as f:
                        saved_meta = f.read().strip().split("|")
                    if len(saved_meta) == 2:
                        saved_last_mod, saved_size_str = saved_meta[0], saved_meta[1]
                        if saved_last_mod == remote_last_mod and int(saved_size_str) == remote_size:
                            use_cached = True
                except Exception:
                    pass
            
            if use_cached:
                return pd.ExcelFile(cache_file)
                
            buf = gc.download_excel("consolidado_personal")
            xl_cloud = pd.ExcelFile(buf)
            if "BD PERSONAL" in xl_cloud.sheet_names or "PERSONAL" in xl_cloud.sheet_names:
                try:
                    buf.seek(0)
                    with open(cache_file, "wb") as f:
                        f.write(buf.getvalue())
                    with open(meta_file, "w", encoding="utf-8") as f:
                        f.write(f"{remote_last_mod}|{remote_size}")
                except Exception as cache_err:
                    print(f"Error escribiendo caché local de personal: {cache_err}")
                buf.seek(0)
                st.session_state.pop("super_load_error", None)
                return pd.ExcelFile(buf)
            else:
                st.session_state["super_load_error"] = f"El archivo cargado desde SharePoint no contiene la hoja 'BD PERSONAL'. Hojas encontradas: {xl_cloud.sheet_names}"
        except Exception as e:
            st.session_state["super_load_error"] = f"Error al conectar con SharePoint: {e}"
            if not force_refresh and os.path.exists(cache_file):
                try:
                    return pd.ExcelFile(cache_file)
                except Exception:
                    pass
            candidate_paths = [
                r"C:\Users\JuanJoseOsorioMolina\U.T SAN VICENTE CES\CENTRAL DE NOVEDADES - Documentos\CONSOLIDADOS\CONSOLIDADO 2026\CONSOLIDADO 2026.xlsx",
                os.path.join(os.path.dirname(excel_path), "CONSOLIDADO 2026.xlsx"),
                "CONSOLIDADO 2026.xlsx",
                r"C:\Users\JuanJoseOsorioMolina\OneDrive - U.T SAN VICENTE CES\CENTRAL DE NOVEDADES CONSOLIDADOS - Documentos\CONSOLIDADOS\CONSOLIDADO 2026\CONSOLIDADO 2026.xlsx",
                r"C:\Users\JuanJoseOsorioMolina\OneDrive - U.T SAN VICENTE CES\CONSOLIDADO 2026.xlsx"
            ]
            for p in candidate_paths:
                if os.path.exists(p):
                    try:
                        return pd.ExcelFile(p)
                    except Exception:
                        try:
                            tmp_p = f"temp_read_{os.path.basename(p)}"
                            import shutil
                            shutil.copy2(p, tmp_p)
                            return pd.ExcelFile(tmp_p)
                        except Exception:
                            continue
            return _open_master_excel(excel_path)
    else:
        candidate_paths = [
            r"C:\Users\JuanJoseOsorioMolina\U.T SAN VICENTE CES\CENTRAL DE NOVEDADES - Documentos\CONSOLIDADOS\CONSOLIDADO 2026\CONSOLIDADO 2026.xlsx",
            os.path.join(os.path.dirname(excel_path), "CONSOLIDADO 2026.xlsx"),
            "CONSOLIDADO 2026.xlsx",
            r"C:\Users\JuanJoseOsorioMolina\OneDrive - U.T SAN VICENTE CES\CENTRAL DE NOVEDADES CONSOLIDADOS - Documentos\CONSOLIDADOS\CONSOLIDADO 2026\CONSOLIDADO 2026.xlsx",
            r"C:\Users\JuanJoseOsorioMolina\OneDrive - U.T SAN VICENTE CES\CONSOLIDADO 2026.xlsx"
        ]
        for p in candidate_paths:
            if os.path.exists(p):
                try:
                    return pd.ExcelFile(p)
                except Exception:
                    try:
                        tmp_p = f"temp_read_{os.path.basename(p)}"
                        import shutil
                        shutil.copy2(p, tmp_p)
                        return pd.ExcelFile(tmp_p)
                    except Exception:
                        continue
        return _open_master_excel(excel_path)


@st.cache_data(ttl=15, show_spinner=False)
def load_supernumeraries(excel_path):
    """
    Carga el personal directamente de la hoja 'BD PERSONAL' de CONSOLIDADO 2026.xlsx.
    Filtra únicamente los médicos con Cargo que contenga 'SUPERNUMERARIO' y Sede 'SUPERNUMERARIO'.
    Si el cargo cambia a 'Medico General' o la sede ya no es 'Supernumerario', la persona se excluye.
    """
    try:
        xl = _open_consolidado_personal(excel_path)
        sheet_target = "BD PERSONAL" if "BD PERSONAL" in xl.sheet_names else ("PERSONAL" if "PERSONAL" in xl.sheet_names else xl.sheet_names[0])
        df = pd.read_excel(xl, sheet_name=sheet_target)
        
        # Mapeo y estandarización de columnas
        col_map = {}
        for c in df.columns:
            c_str = str(c).strip().upper()
            if c_str in ("CEDULA", "DOCUMENTO", "IDENTIFICACION", "ID", "CÉDULA"):
                col_map[c] = "CEDULA"
            elif c_str in ("NOMBRES Y APELLIDOS", "NOMBRES_Y_APELLIDOS", "NOMBRE COMPLETO", "NOMBRES", "APELLIDOS Y NOMBRES", "PROFESIONAL"):
                col_map[c] = "NOMBRES Y APELLIDOS"
            elif c_str in ("CARGO", "CARGO DE TRABAJO"):
                col_map[c] = "CARGO"
            elif c_str in ("SEDE / CECO", "SEDE", "CECO", "SEDE_CECO"):
                col_map[c] = "SEDE / CECO"
            elif c_str in ("CORREO", "EMAIL", "CORREO ELECTRONICO", "CORREO ELECTRÓNICO"):
                col_map[c] = "CORREO"
            elif c_str in ("CELULAR", "TELEFONO", "MOVIL"):
                col_map[c] = "CELULAR"
            elif c_str in ("STATUS", "ESTADO", "ESTADO OPERATIVO"):
                col_map[c] = "STATUS"
            elif c_str in ("OBSERVACIONES", "OBSERVACION", "NOTAS"):
                col_map[c] = "OBSERVACIONES"

        df = df.rename(columns=col_map)
        
        for req_col in ["CEDULA", "NOMBRES Y APELLIDOS", "CARGO", "SEDE / CECO", "CORREO", "STATUS", "CELULAR", "OBSERVACIONES"]:
            if req_col not in df.columns:
                df[req_col] = ""

        # Limpieza de textos y valores nulos
        df["CARGO"] = df["CARGO"].fillna("").astype(str).str.strip()
        df["SEDE / CECO"] = df["SEDE / CECO"].fillna("").astype(str).str.strip()
        df["STATUS"] = df["STATUS"].fillna("SI").astype(str).str.strip()
        df["CORREO"] = df["CORREO"].fillna("").astype(str).str.strip()

        # Regla de filtrado estricta:
        # Traer únicamente médicos cuyo Cargo contenga 'SUPERNUMERARIO' (ej: Medico General Supernumerario), ignorando el filtro por Sede.
        is_super_cargo = df["CARGO"].str.upper().str.contains("SUPERNUMERARI", na=False)
        
        # Fallback solo si la columna CARGO estaba totalmente vacía (ej. archivos locales antiguos)
        if not is_super_cargo.any():
            is_super_doc = df["SEDE / CECO"].str.upper().str.contains("SUPERNUMERARI", na=False)
        else:
            is_super_doc = is_super_cargo
        
        status_upper = df["STATUS"].str.upper()
        is_active_status = status_upper.isin(["ACTIVO", "SI", "YES", "1", ""]) | ~status_upper.isin(["INACTIVO", "NO", "RETIRADO", "EGRESADO", "BAJA", "DESACTIVADO"])

        df_super = df[is_super_doc & is_active_status].copy()

        df_super["CEDULA"] = df_super["CEDULA"].apply(
            lambda x: str(int(x)) if pd.notna(x) and str(x).replace(".0", "").isdigit() else str(x).strip()
        )
        df_super["NOMBRES Y APELLIDOS"] = (
            df_super["NOMBRES Y APELLIDOS"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .apply(lambda x: re.sub(r"\s+", " ", x))
        )
        df_super["CORREO"] = df_super["CORREO"].fillna("").astype(str).str.strip()
        df_super["OBSERVACIONES"] = df_super["OBSERVACIONES"].fillna("").astype(str).str.strip()

        # Crear nombres de columna amigables coincidentes con el Excel original
        df_super["Cédula"] = df_super["CEDULA"]
        df_super["Sede"] = df_super["SEDE / CECO"]
        df_super["Cargo"] = df_super["CARGO"]
        df_super["Profesional"] = df_super["NOMBRES Y APELLIDOS"]
        df_super["Estado"] = df_super["STATUS"]
        df_super["Correo"] = df_super["CORREO"]

        # Combinar médicos agregados manualmente por contingencia desde PERSONAL_MANUAL
        try:
            df_manual = load_manual_supernumeraries(excel_path)
            if not df_manual.empty:
                existing_cedulas = set(df_super["CEDULA"].astype(str).tolist())
                existing_names = set(df_super["NOMBRES Y APELLIDOS"].astype(str).tolist())
                manual_rows = []
                for _, m_row in df_manual.iterrows():
                    m_ced = str(m_row["CEDULA"]).strip()
                    m_name = str(m_row["NOMBRES Y APELLIDOS"]).strip().upper()
                    if m_ced not in existing_cedulas and m_name not in existing_names:
                        manual_rows.append({
                            "CEDULA": m_ced,
                            "NOMBRES Y APELLIDOS": m_name,
                            "CARGO": str(m_row.get("CARGO", "MEDICO GENERAL SUPERNUMERARIO")).strip(),
                            "SEDE / CECO": str(m_row.get("SEDE / CECO", "SUPERNUMERARIOS")).strip(),
                            "CORREO": "",
                            "STATUS": "SI",
                            "CELULAR": str(m_row.get("CELULAR", "")).strip(),
                            "OBSERVACIONES": str(m_row.get("OBSERVACIONES", "Registro manual contingencia")).strip(),
                            "Cédula": m_ced,
                            "Sede": str(m_row.get("SEDE / CECO", "SUPERNUMERARIOS")).strip(),
                            "Cargo": str(m_row.get("CARGO", "MEDICO GENERAL SUPERNUMERARIO")).strip(),
                            "Profesional": m_name,
                            "Estado": "SI",
                            "Correo": ""
                        })
                if manual_rows:
                    df_super = pd.concat([df_super, pd.DataFrame(manual_rows)], ignore_index=True)
        except Exception as manual_err:
            print(f"Error combinando médicos manuales de contingencia: {manual_err}")
        
        return (
            df_super[["Cédula", "Sede", "Cargo", "Profesional", "Estado", "Correo", "CEDULA", "NOMBRES Y APELLIDOS", "CARGO", "SEDE / CECO", "STATUS", "CORREO", "CELULAR", "OBSERVACIONES"]]
            .sort_values(by="Profesional")
            .reset_index(drop=True)
        )
    except Exception as e:
        print(f"Error cargando supernumerarios de BD PERSONAL: {e}")
        return pd.DataFrame(columns=["Cédula", "Sede", "Cargo", "Profesional", "Estado", "Correo", "CEDULA", "NOMBRES Y APELLIDOS", "CARGO", "SEDE / CECO", "STATUS", "CORREO", "CELULAR", "OBSERVACIONES"])


# ---------------------------------------------------------------------------
# Operaciones de escritura (wrapper sin cambios de firma)
# ---------------------------------------------------------------------------

def update_shift_cell(excel_path, sheet_name, row_idx, col_idx, new_name,
                      date_val=None, observation='', original_name=None,
                      clasificacion='Secuencia Normal'):
    if not date_val:
        df_shifts, _ = load_data(excel_path)
        match = df_shifts[(df_shifts['Sheet'] == sheet_name) &
                          (df_shifts['Excel_Row'] == row_idx) &
                          (df_shifts['Excel_Col'] == col_idx)]
        if not match.empty:
            date_val  = match.iloc[0]['Date']
            orig_name = match.iloc[0]['Supernumerary']
        else:
            date_val  = datetime.date.today()
            orig_name = ""
    else:
        if original_name:
            orig_name = original_name
        else:
            df_shifts, _ = load_data(excel_path)
            match = df_shifts[(df_shifts['Sheet'] == sheet_name) &
                              (df_shifts['Excel_Row'] == row_idx) &
                              (df_shifts['Excel_Col'] == col_idx)]
            orig_name = match.iloc[0]['Supernumerary'] if not match.empty else ""

    if original_name:
        orig_name = original_name

    cleaned_new = str(new_name).strip().upper() if new_name else ""
    mod_data = {
        'sheet': sheet_name, 'date': date_val,
        'original_name': orig_name, 'new_name': cleaned_new,
        'row': row_idx, 'col': col_idx,
        'type': 'ELIMINAR' if not cleaned_new else 'REEMPLAZAR',
        'observaciones': observation, 'clasificacion': clasificacion
    }
    save_modification(excel_path, mod_data)
    return True


def delete_shift_cell(excel_path, sheet_name, row_idx, col_idx,
                      date_val=None, observation='', original_name=None,
                      clasificacion='Secuencia Normal'):
    return update_shift_cell(excel_path, sheet_name, row_idx, col_idx, None,
                             date_val, observation, original_name, clasificacion)


def add_shift_to_date(excel_path, sheet_name, target_date, supernumerary_name,
                      observation='', clasificacion='Secuencia Normal'):
    new_name = str(supernumerary_name).strip().upper()
    if not new_name:
        raise ValueError("El nombre del supernumerario no puede estar vacío.")
    mod_data = {
        'sheet': sheet_name, 'date': target_date,
        'original_name': '', 'new_name': new_name,
        'row': 0, 'col': 0, 'type': 'AGREGAR',
        'observaciones': observation, 'clasificacion': clasificacion
    }
    save_modification(excel_path, mod_data)
    return True


def duplicate_schedule_batch(excel_path, target_sheet, target_date, shifts_to_delete_list, shifts_to_add_list):
    mods_list = []
    
    # 1. Preparar las eliminaciones
    for s_del in shifts_to_delete_list:
        mods_list.append({
            'sheet': s_del['sheet'],
            'date': target_date,
            'original_name': s_del['doctor'],
            'new_name': '',
            'row': s_del['row'],
            'col': s_del['col'],
            'type': 'ELIMINAR',
            'observaciones': '',
            'clasificacion': 'Secuencia Normal'
        })
        
    # 2. Preparar las adiciones
    for s_add in shifts_to_add_list:
        mods_list.append({
            'sheet': target_sheet,
            'date': target_date,
            'original_name': '',
            'new_name': s_add['doctor'],
            'row': 0,
            'col': 0,
            'type': 'AGREGAR',
            'observaciones': s_add.get('observation', ''),
            'clasificacion': s_add.get('classification', 'Secuencia Normal')
        })
        
    if mods_list:
        save_modifications_batch(excel_path, mods_list)
    return True


def add_shifts_batch(excel_path, shifts_list):
    mods_list = []
    for s in shifts_list:
        new_name = str(s['doc']).strip().upper()
        if not new_name:
            continue
        mods_list.append({
            'sheet': s['sheet'],
            'date': s['date'],
            'original_name': '',
            'new_name': new_name,
            'row': 0,
            'col': 0,
            'type': 'AGREGAR',
            'observaciones': s.get('obs', ''),
            'clasificacion': s.get('clasificacion', 'Secuencia Normal')
        })
    if mods_list:
        save_modifications_batch(excel_path, mods_list)
    return True

